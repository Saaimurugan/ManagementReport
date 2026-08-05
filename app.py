"""
Flask UI for JIRA Dashboard Generator
--------------------------------------
1. Upload a JIRA Excel file
2. Convert it to jira_data.json  (logic from generate_dashboard_data.py)
3. Embed the JSON into dashboard_standalone.html (logic from create_standalone_dashboard.py)
4. Serve the finished dashboard for download / inline preview
""" 

import json
import os
import re
import traceback
import logging
from io import BytesIO

import pandas as pd
from flask import (
    Flask,
    Response,
    flash,
    redirect,
    render_template_string,
    request,
    send_file,
    url_for,
)
from werkzeug.utils import secure_filename

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

try:
    from create_template import build_template
except ImportError as e:
    logger.error(f"Failed to import create_template: {e}")
    build_template = None

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ALLOWED_EXTENSIONS = {"xlsx", "xls"}

app = Flask(__name__)
app.secret_key = "jira-dashboard-secret-key-change-me"
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# Note: In serverless environment, we process everything in memory
# No file uploads to disk are performed

# ---------------------------------------------------------------------------
# HTML template (inline – no separate templates folder needed)
# ---------------------------------------------------------------------------
PAGE_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>JIRA Dashboard Generator</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    body {
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
      background: #090c10;
      color: #c9d1d9;
      min-height: 100vh;
      display: grid;
      place-items: center;
      padding: 2rem 1rem;
    }

    /* ── Page shell ─────────────────────────────────────────────────── */
    .shell {
      width: 100%;
      max-width: 680px;
      display: flex;
      flex-direction: column;
      gap: 0;
    }

    /* ── Top header bar ─────────────────────────────────────────────── */
    .top-bar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      padding: 1.1rem 1.75rem;
      background: #161b22;
      border: 1px solid #30363d;
      border-bottom: none;
      border-radius: 14px 14px 0 0;
    }
    .brand { display: flex; align-items: center; gap: .75rem; }
    .brand-icon {
      width: 38px; height: 38px; border-radius: 9px;
      background: linear-gradient(135deg, #1f6feb 0%, #388bfd 100%);
      display: flex; align-items: center; justify-content: center;
      box-shadow: 0 2px 8px rgba(31,111,235,.45);
      flex-shrink: 0;
    }
    .brand h1 { font-size: 1.15rem; font-weight: 700; color: #e6edf3; letter-spacing: -.01em; }
    .brand p  { font-size: .78rem; color: #6e7681; margin-top: 1px; }
    .badge {
      font-size: .7rem; font-weight: 600; padding: .25rem .6rem;
      background: #1c2333; border: 1px solid #30363d;
      border-radius: 20px; color: #58a6ff;
    }

    /* ── Step tracker ───────────────────────────────────────────────── */
    .stepper {
      display: flex;
      align-items: center;
      padding: 1.1rem 1.75rem;
      background: #0d1117;
      border-left: 1px solid #30363d;
      border-right: 1px solid #30363d;
      gap: 0;
    }
    .step-item {
      display: flex;
      align-items: center;
      flex: 1;
      gap: 0;
    }
    .step-item:last-child { flex: 0; }
    .step-content {
      display: flex;
      flex-direction: column;
      align-items: center;
      gap: .3rem;
      min-width: 64px;
    }
    .step-dot {
      width: 32px; height: 32px; border-radius: 50%;
      background: #161b22; border: 2px solid #30363d;
      display: flex; align-items: center; justify-content: center;
      font-size: .78rem; font-weight: 700; color: #484f58;
      transition: all .25s;
    }
    .step-item.active .step-dot {
      background: #1f6feb; border-color: #388bfd;
      color: #fff; box-shadow: 0 0 0 4px rgba(31,111,235,.2);
    }
    .step-item.done .step-dot {
      background: #238636; border-color: #2ea043; color: #fff;
    }
    .step-label {
      font-size: .7rem; font-weight: 600; color: #484f58;
      white-space: nowrap; letter-spacing: .02em; text-transform: uppercase;
    }
    .step-item.active .step-label { color: #58a6ff; }
    .step-item.done  .step-label  { color: #3fb950; }
    .step-connector {
      flex: 1; height: 2px;
      background: #21262d;
      margin: 0 .5rem;
      margin-bottom: 1.4rem;
      transition: background .25s;
    }
    .step-connector.done { background: #2ea043; }

    /* ── Main card body ─────────────────────────────────────────────── */
    .card-body {
      background: #161b22;
      border: 1px solid #30363d;
      border-top: none;
      border-radius: 0 0 14px 14px;
      padding: 2rem 1.75rem 1.75rem;
      box-shadow: 0 12px 40px rgba(0,0,0,.5);
    }

    /* ── Upload zone ─────────────────────────────────────────────────── */
    label.upload-area {
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: center;
      gap: .55rem;
      border: 2px dashed #30363d;
      border-radius: 10px;
      padding: 2.5rem 1.5rem;
      cursor: pointer;
      transition: border-color .2s, background .2s;
      margin-bottom: 1.25rem;
      background: #0d1117;
    }
    label.upload-area:hover,
    label.upload-area.drag-over { border-color: #388bfd; background: #0c1929; }
    .upload-icon {
      width: 52px; height: 52px; border-radius: 12px;
      background: #1c2333; border: 1px solid #30363d;
      display: flex; align-items: center; justify-content: center;
      font-size: 1.5rem;
    }
    .upload-title { font-size: .95rem; font-weight: 600; color: #e6edf3; }
    .upload-sub   { font-size: .8rem; color: #6e7681; }
    .file-chosen  { font-size: .85rem; color: #58a6ff; word-break: break-all; text-align: center; }
    #file-input   { display: none; }

    /* ── Buttons ─────────────────────────────────────────────────────── */
    .btn {
      display: flex; align-items: center; justify-content: center;
      gap: .5rem; width: 100%; padding: .75rem 1rem;
      border: none; border-radius: 8px;
      font-size: .95rem; font-weight: 600; cursor: pointer;
      transition: filter .15s, transform .1s;
      text-decoration: none;
    }
    .btn:active { transform: scale(.98); }
    .btn-primary { background: linear-gradient(135deg,#1f6feb,#388bfd); color: #fff; }
    .btn-primary:hover { filter: brightness(1.12); }
    .btn-primary:disabled { opacity: .45; cursor: not-allowed; filter: none; }
    .btn-success { background: linear-gradient(135deg,#238636,#2ea043); color: #fff; }
    .btn-success:hover { filter: brightness(1.12); }
    .btn-ghost {
      background: #21262d; color: #8b949e;
      border: 1px solid #30363d;
    }
    .btn-ghost:hover { background: #292e36; color: #c9d1d9; }
    .btn-outline {
      background: transparent; color: #58a6ff;
      border: 1px solid #388bfd;
    }
    .btn-outline:hover { background: rgba(56,139,253,.1); }
    .btn-row { display: flex; gap: .75rem; margin-top: .75rem; }
    .btn-row .btn { flex: 1; }

    /* ── Alerts ──────────────────────────────────────────────────────── */
    .alert {
      border-radius: 8px; padding: .85rem 1rem;
      margin-bottom: 1.25rem; font-size: .875rem;
      display: flex; align-items: flex-start; gap: .6rem;
      white-space: pre-line; /* Allow line breaks in error messages */
      line-height: 1.4;
    }
    .alert-danger  { background: #2d1517; border: 1px solid #6e2428; color: #ff7b72; }
    .alert-success { background: #0d2d1a; border: 1px solid #1a4731; color: #3fb950; }
    .alert-info    { background: #0c1929; border: 1px solid #1b3152; color: #79c0ff; }
    .alert span:first-child { flex-shrink: 0; margin-top: 2px; } /* Icon positioning */

    /* ── Stats row (step 2) ──────────────────────────────────────────── */
    .stat-card {
      background: #0d1117; border: 1px solid #30363d;
      border-radius: 10px; padding: 1.1rem 1.4rem;
      margin-bottom: 1.25rem;
      display: flex; align-items: center; gap: 1rem;
    }
    .stat-icon {
      font-size: 1.6rem; width: 48px; height: 48px;
      background: #161b22; border-radius: 10px;
      display: flex; align-items: center; justify-content: center;
      flex-shrink: 0; border: 1px solid #30363d;
    }
    .stat-label { font-size: .75rem; color: #6e7681; text-transform: uppercase;
                  letter-spacing: .06em; font-weight: 600; }
    .stat-value { font-size: 1.5rem; font-weight: 700; color: #e6edf3;
                  line-height: 1.1; margin-top: .1rem; }
    .stat-sub   { font-size: .78rem; color: #3fb950; margin-top: .15rem; }

    /* ── Result box (step 3) ─────────────────────────────────────────── */
    .result-box {
      background: #0d1117; border: 1px solid #1a4731;
      border-radius: 10px; padding: 1.25rem 1.4rem;
      margin-bottom: 1.25rem;
    }
    .result-box h3 { font-size: .95rem; font-weight: 700; color: #e6edf3; margin-bottom: .4rem; }
    .result-box p  { font-size: .83rem; color: #6e7681; margin-bottom: 1.1rem; line-height: 1.5; }

    /* ── Progress bar ────────────────────────────────────────────────── */
    .progress-wrap { display: none; margin-bottom: 1.1rem; }
    .progress-track {
      height: 5px; border-radius: 3px;
      background: #21262d; overflow: hidden;
    }
    .progress-fill {
      height: 100%; width: 0%;
      background: linear-gradient(90deg,#1f6feb,#58a6ff);
      border-radius: 3px; transition: width .45s ease;
    }
    .progress-label { font-size: .75rem; color: #6e7681; margin-top: .4rem; }

    /* ── Template hint ───────────────────────────────────────────────── */
    .template-hint {
      display: flex; align-items: center; justify-content: center;
      gap: .4rem; margin-top: 1rem;
      font-size: .8rem; color: #6e7681;
    }
    .template-hint a {
      color: #58a6ff; text-decoration: none; font-weight: 500;
      display: inline-flex; align-items: center; gap: .25rem;
    }
    .template-hint a:hover { text-decoration: underline; }

    /* ── Divider ─────────────────────────────────────────────────────── */
    .divider {
      border: none; border-top: 1px solid #21262d;
      margin: 1.25rem 0;
    }

    /* ── Footer ──────────────────────────────────────────────────────── */
    footer {
      text-align: center; margin-top: 1.25rem;
      font-size: .72rem; color: #30363d; letter-spacing: .03em;
    }
  </style>
</head>
<body>
<div class="shell">

  <!-- ── Top bar ── -->
  <div class="top-bar">
    <div class="brand">
      <div class="brand-icon">
        <svg width="22" height="22" viewBox="0 0 24 24" fill="none">
          <path d="M3 3h7v7H3zM14 3h7v7h-7zM14 14h7v7h-7zM3 14h7v7H3z"
                stroke="#fff" stroke-width="1.8" stroke-linejoin="round"/>
        </svg>
      </div>
      <div>
        <h1>JIRA Dashboard Generator</h1>
        <p>Upload Excel → Convert → Generate Standalone Report</p>
      </div>
    </div>
    <span class="badge">v1.0</span>
  </div>

  <!-- ── Step tracker ── -->
  <div class="stepper">
    <div class="step-item {{ 'done' if step > 1 else 'active' if step == 1 else '' }}">
      <div class="step-content">
        <div class="step-dot">{% if step > 1 %}✓{% else %}1{% endif %}</div>
        <div class="step-label">Upload</div>
      </div>
    </div>
    <div class="step-connector {{ 'done' if step > 1 else '' }}"></div>
    <div class="step-item {{ 'done' if step > 2 else 'active' if step == 2 else '' }}">
      <div class="step-content">
        <div class="step-dot">{% if step > 2 %}✓{% else %}2{% endif %}</div>
        <div class="step-label">Convert</div>
      </div>
    </div>
    <div class="step-connector {{ 'done' if step > 2 else '' }}"></div>
    <div class="step-item {{ 'done' if step > 3 else 'active' if step == 3 else '' }}">
      <div class="step-content">
        <div class="step-dot">{% if step > 3 %}✓{% else %}3{% endif %}</div>
        <div class="step-label">Generate</div>
      </div>
    </div>
  </div>

  <!-- ── Card body ── -->
  <div class="card-body">

    <!-- Alerts -->
    {% for category, message in messages %}
    <div class="alert alert-{{ category }}">
      <span>{% if category == 'danger' %}⚠{% elif category == 'success' %}✓{% else %}ℹ{% endif %}</span>
      <span>{{ message }}</span>
    </div>
    {% endfor %}

    {% if step == 1 %}
    <!-- STEP 1: Upload -->
    <form id="upload-form" method="post" action="{{ url_for('upload') }}" enctype="multipart/form-data">
      <label class="upload-area" for="file-input" id="drop-zone">
        <div class="upload-icon">�</div>
        <span class="upload-title" id="chosen-file">Drop your JIRA Excel file here</span>
        <span class="upload-sub">or click to browse &nbsp;·&nbsp; .xlsx / .xls</span>
      </label>
      <input id="file-input" type="file" name="excel_file" accept=".xlsx,.xls" required />

      <div class="progress-wrap" id="progress-wrap">
        <div class="progress-track"><div class="progress-fill" id="progress-fill"></div></div>
        <div class="progress-label" id="progress-label">Uploading…</div>
      </div>

      <button type="submit" class="btn btn-primary" id="submit-btn">
        <svg width="16" height="16" viewBox="0 0 16 16" fill="currentColor">
          <path d="M8 15a.75.75 0 0 1-.75-.75V8.75H4.53a.25.25 0 0 1-.177-.427l3.396-3.396a.25.25 0 0 1 .354 0l3.396 3.396a.25.25 0 0 1-.177.427H8.75v5.5A.75.75 0 0 1 8 15zM1.5 4a.75.75 0 0 1 .75-.75h11.5a.75.75 0 0 1 0 1.5H2.25A.75.75 0 0 1 1.5 4z"/>
        </svg>
        Upload &amp; Convert to JSON
      </button>
    </form>

    <div class="template-hint">
      <span>New to this?</span>
      <a href="{{ url_for('download_template') }}">
        <svg width="13" height="13" viewBox="0 0 16 16" fill="currentColor">
          <path d="M8 1a.75.75 0 0 1 .75.75v5.5h2.72a.25.25 0 0 1 .177.427l-3.396 3.396a.25.25 0 0 1-.353 0L4.503 7.677A.25.25 0 0 1 4.68 7.25H7.25V1.75A.75.75 0 0 1 8 1zM1.5 12a.75.75 0 0 1 .75-.75h11.5a.75.75 0 0 1 0 1.5H2.25A.75.75 0 0 1 1.5 12z"/>
        </svg>
        Download sample Excel template
      </a>
    </div>

    {% elif step == 2 %}
    <!-- STEP 2: Convert done, generate dashboard -->
    <div class="stat-card">
      <div class="stat-icon">🗂️</div>
      <div>
        <div class="stat-label">Records converted</div>
        <div class="stat-value">{{ record_count }}</div>
        <div class="stat-sub">✓ jira_data.json saved</div>
      </div>
    </div>

    <form method="post" action="{{ url_for('generate') }}">
      <button type="submit" class="btn btn-primary">
        <svg width="16" height="16" viewBox="0 0 16 16" fill="currentColor">
          <path d="M2 2.5A2.5 2.5 0 0 1 4.5 0h7A2.5 2.5 0 0 1 14 2.5v10.042a.75.75 0 0 1-1.225.575l-2.275-1.74-2.275 1.74a.75.75 0 0 1-.9 0L5.05 11.377l-2.275 1.74A.75.75 0 0 1 1.5 12.5V2.5zm2.5-1A1 1 0 0 0 3.5 2.5v8.653l1.525-1.166a.75.75 0 0 1 .9 0L8.2 11.727l2.275-1.74a.75.75 0 0 1 .9 0l1.525 1.166V2.5a1 1 0 0 0-1-1h-7z"/>
        </svg>
        Generate Standalone Dashboard
      </button>
    </form>

    <hr class="divider" />
    <form method="get" action="{{ url_for('index') }}">
      <button type="submit" class="btn btn-ghost">← Upload a different file</button>
    </form>

    {% elif step == 3 %}
    <!-- STEP 3: Done -->
    <div class="result-box">
      <h3>🎉 Dashboard ready</h3>
      <p>Your self-contained HTML report is ready. Download it and open in any browser — no server or separate JSON file needed.</p>
      <a href="{{ url_for('download') }}" class="btn btn-success" style="text-decoration:none">
        <svg width="15" height="15" viewBox="0 0 16 16" fill="currentColor">
          <path d="M8 1a.75.75 0 0 1 .75.75v5.5h2.72a.25.25 0 0 1 .177.427l-3.396 3.396a.25.25 0 0 1-.353 0L4.503 7.677A.25.25 0 0 1 4.68 7.25H7.25V1.75A.75.75 0 0 1 8 1zM1.5 12a.75.75 0 0 1 .75-.75h11.5a.75.75 0 0 1 0 1.5H2.25A.75.75 0 0 1 1.5 12z"/>
        </svg>
        Download dashboard_standalone.html
      </a>
      <div class="btn-row">
        <a href="{{ url_for('preview') }}" target="_blank" class="btn btn-outline" style="text-decoration:none">
          🔍 Preview in browser
        </a>
      </div>
    </div>

    <form method="get" action="{{ url_for('index') }}">
      <button type="submit" class="btn btn-ghost">← Start over</button>
    </form>
    {% endif %}

  </div><!-- /card-body -->
</div><!-- /shell -->

<footer>JIRA Management Report &nbsp;·&nbsp; Flask UI</footer>

<script>
  // File picker feedback
  const fileInput = document.getElementById('file-input');
  const fileLabel = document.getElementById('chosen-file');
  if (fileInput) {
    fileInput.addEventListener('change', () => {
      if (fileInput.files.length) {
        fileLabel.textContent = fileInput.files[0].name;
        fileLabel.style.color = '#58a6ff';
      } else {
        fileLabel.textContent = 'Drop your JIRA Excel file here';
        fileLabel.style.color = '';
      }
    });
  }

  // Drag-and-drop
  const dropZone = document.getElementById('drop-zone');
  if (dropZone) {
    ['dragenter', 'dragover'].forEach(evt =>
      dropZone.addEventListener(evt, e => {
        e.preventDefault();
        dropZone.classList.add('drag-over');
      })
    );
    ['dragleave', 'drop'].forEach(evt =>
      dropZone.addEventListener(evt, e => {
        e.preventDefault();
        dropZone.classList.remove('drag-over');
      })
    );
    dropZone.addEventListener('drop', e => {
      const files = e.dataTransfer.files;
      if (files.length && fileInput) {
        try { fileInput.files = files; } catch(_) {}
        fileLabel.textContent = files[0].name;
        fileLabel.style.color = '#58a6ff';
      }
    });
  }

  // Upload progress animation
  const uploadForm = document.getElementById('upload-form');
  if (uploadForm) {
    uploadForm.addEventListener('submit', () => {
      const wrap = document.getElementById('progress-wrap');
      const fill = document.getElementById('progress-fill');
      const lbl  = document.getElementById('progress-label');
      const btn  = document.getElementById('submit-btn');
      if (wrap) wrap.style.display = 'block';
      if (btn)  btn.disabled = true;
      [
        [350,  25, 'Uploading file…'],
        [800,  55, 'Converting to JSON…'],
        [1300, 80, 'Almost done…'],
      ].forEach(([delay, pct, text]) => {
        setTimeout(() => {
          if (fill) fill.style.width = pct + '%';
          if (lbl)  lbl.textContent  = text;
        }, delay);
      });
    });
  }
</script>
</body>
</html>
"""

# ---------------------------------------------------------------------------
# Core processing helpers
# ---------------------------------------------------------------------------

def validate_excel_format(df) -> tuple[bool, list]:
    """
    Validate Excel file format and return validation status and error messages.
    
    Returns:
        tuple: (is_valid, list_of_error_messages)
    """
    errors = []
    
    # Define expected columns and their types
    expected_columns = {
        "Client": "string",
        "Type": "string",
        "Project": "string", 
        "Sprint": "string",
        "EPIC": "string",
        "Story": "string",
        "Task": "string",
        "Quality": "numeric",
        "Budget Planned": "numeric",
        "Budget Consumed": "numeric", 
        "Budget Remaining": "numeric",
        "Saving Planned": "numeric",
        "Saving Achived": "numeric",  # Note: keeping original spelling from data
        "Saving Pending": "numeric",
        "Start Date": "date",
        "Date": "date",
        "Story Points": "numeric",
        "Priority": "string",
        "Assignee": "string", 
        "Status": "string",
        "Dependencies": "string"
    }
    
    # Check if file is empty
    if df.empty:
        errors.append("The Excel file is empty. Please provide a file with data.")
        return False, errors
    
    # Check for required columns
    missing_columns = []
    for col in expected_columns.keys():
        if col not in df.columns:
            missing_columns.append(col)
    
    if missing_columns:
        errors.append(f"Missing required columns: {', '.join(missing_columns)}")
    
    # Check for extra unexpected columns (warn but don't fail)
    extra_columns = [col for col in df.columns if col not in expected_columns and col != "Cost"]
    if extra_columns:
        errors.append(f"Warning: Unexpected columns found (will be ignored): {', '.join(extra_columns)}")
    
    # Validate data types for existing columns
    for col, expected_type in expected_columns.items():
        if col not in df.columns:
            continue
            
        # Get non-null values for validation
        non_null_values = df[col].dropna()
        if len(non_null_values) == 0:
            continue  # Skip validation if all values are null
            
        try:
            if expected_type == "numeric":
                # Check if values can be converted to numeric
                pd.to_numeric(non_null_values, errors='raise')
                
            elif expected_type == "date":
                # Check if values can be converted to datetime
                pd.to_datetime(non_null_values, errors='raise')
                
            elif expected_type == "string":
                # For string columns, check for reasonable length
                max_length = non_null_values.astype(str).str.len().max()
                if max_length > 500:
                    errors.append(f"Column '{col}' contains values that are too long (max: {max_length} chars). Please keep text under 500 characters.")
                    
        except (ValueError, TypeError) as e:
            errors.append(f"Column '{col}' contains invalid {expected_type} values. Please check the data format.")
    
    # Validate specific business rules
    
    # Check Priority values
    if "Priority" in df.columns:
        valid_priorities = ["Low", "Medium", "High", "Critical", "Normal"]
        invalid_priorities = df[df["Priority"].notna()]["Priority"].unique()
        invalid_priorities = [p for p in invalid_priorities if str(p).strip() not in valid_priorities]
        if invalid_priorities:
            errors.append(f"Invalid Priority values found: {invalid_priorities}. Valid values are: {valid_priorities}")
    
    # Check Status values
    if "Status" in df.columns:
        valid_statuses = ["To Do", "In Progress", "Done", "Blocked", "Review"]
        invalid_statuses = df[df["Status"].notna()]["Status"].unique()
        invalid_statuses = [s for s in invalid_statuses if str(s).strip() not in valid_statuses]
        if invalid_statuses:
            errors.append(f"Invalid Status values found: {invalid_statuses}. Valid values are: {valid_statuses}")
    
    # Check Type values
    if "Type" in df.columns:
        valid_types = ["Epic", "Story", "Task", "Bug", "Subtask"]
        invalid_types = df[df["Type"].notna()]["Type"].unique() 
        invalid_types = [t for t in invalid_types if str(t).strip() not in valid_types]
        if invalid_types:
            errors.append(f"Invalid Type values found: {invalid_types}. Valid values are: {valid_types}")
    
    # Check for negative budget values
    budget_columns = ["Budget Planned", "Budget Consumed", "Budget Remaining", "Saving Planned", "Saving Achived", "Saving Pending"]
    for col in budget_columns:
        if col in df.columns:
            negative_values = df[df[col] < 0][col].notna()
            if len(negative_values) > 0:
                errors.append(f"Column '{col}' contains negative values, which may not be valid for budget data.")
    
    # Check Story Points are positive integers
    if "Story Points" in df.columns:
        story_points = df["Story Points"].dropna()
        if len(story_points) > 0:
            try:
                # Check if they're integers and positive
                int_values = story_points.astype(int)
                if (int_values < 0).any():
                    errors.append("Story Points cannot be negative.")
                if (story_points != int_values).any():
                    errors.append("Story Points should be whole numbers (integers).")
            except (ValueError, TypeError):
                errors.append("Story Points must be numeric values.")
    
    # Check date format
    if "Date" in df.columns:
        date_col = df["Date"].dropna()
        if len(date_col) > 0:
            try:
                converted_dates = pd.to_datetime(date_col, errors='raise')
                # Check for future dates beyond reasonable project timeline (2 years from now)
                import datetime
                max_date = datetime.datetime.now() + datetime.timedelta(days=730)
                if (converted_dates > max_date).any():
                    errors.append("Some dates are too far in the future (more than 2 years). Please verify the dates.")
            except:
                errors.append("Date column contains invalid date formats. Please use a standard date format (YYYY-MM-DD, MM/DD/YYYY, etc.).")
    
    # Summary validation
    if not missing_columns and len([e for e in errors if not e.startswith("Warning:")]) == 0:
        return True, errors  # Valid with possible warnings
    else:
        return False, errors


def excel_to_json_memory(file_stream) -> tuple[dict, int]:
    """Process Excel file in memory without saving to disk."""
    try:
        logger.debug("Processing Excel file in memory")
        
        # Read Excel directly from the file stream
        df = pd.read_excel(file_stream)
        logger.debug(f"Excel file loaded with {len(df)} rows and columns: {list(df.columns)}")
        
        # Validate Excel format
        is_valid, validation_errors = validate_excel_format(df)
        
        if not is_valid:
            error_msg = "Excel file validation failed:\n" + "\n".join(validation_errors)
            logger.error(f"Validation failed: {error_msg}")
            raise ValueError(error_msg)
        
        # Log warnings if any
        warnings = [e for e in validation_errors if e.startswith("Warning:")]
        if warnings:
            for warning in warnings:
                logger.warning(warning)
        
        df = df.drop(columns=["Cost"], errors="ignore")
        
        # Handle date conversion more safely
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors='coerce').dt.strftime("%Y-%m-%d")
        
        # Handle Start Date conversion
        if "Start Date" in df.columns:
            df["Start Date"] = pd.to_datetime(df["Start Date"], errors='coerce').dt.strftime("%Y-%m-%d")
        
        # Handle Actual Completion Date conversion
        if "Actual Completion Date" in df.columns:
            df["Actual Completion Date"] = pd.to_datetime(df["Actual Completion Date"], errors='coerce').dt.strftime("%Y-%m-%d")
        
        # Replace NaN values with None for JSON compatibility
        df = df.where(pd.notnull(df), None)
        
        # Convert to dict, ensuring all string values are properly handled
        data = []
        for record in df.to_dict(orient="records"):
            cleaned_record = {}
            for key, value in record.items():
                if value is None:
                    cleaned_record[key] = None
                elif isinstance(value, str):
                    # Ensure strings are properly encoded and handle any problematic characters
                    # Python's json.dumps will automatically handle escaping
                    cleaned_record[key] = str(value)
                else:
                    cleaned_record[key] = value
            data.append(cleaned_record)
        
        logger.info(f"Successfully processed {len(data)} records in memory")
        
        return data, len(data)
    except Exception as e:
        logger.error(f"Error in excel_to_json_memory: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise


def build_standalone_memory(data: list, template_content: str) -> str:
    """Build standalone dashboard in memory without file I/O."""
    try:
        import pandas as pd
        logger.debug("Building standalone dashboard in memory")

        # Clean data for JSON serialization
        cleaned_data = []
        for record in data:
            cleaned_record = {}
            for key, value in record.items():
                if pd.isna(value):
                    cleaned_record[key] = None
                elif isinstance(value, pd.Timestamp):
                    cleaned_record[key] = value.strftime("%Y-%m-%d")
                elif isinstance(value, str):
                    # Properly escape special characters in strings
                    cleaned_record[key] = value
                else:
                    cleaned_record[key] = value
            cleaned_data.append(cleaned_record)

        loader_pattern = re.compile(
            r"        // Load data from jira_data\.json\s*\n"
            r"        fetch\('jira_data\.json'\).*?"
            r"\}\);",
            re.DOTALL,
        )

        # Use ensure_ascii=False to properly handle Unicode characters
        # and let json.dumps handle all escape sequences properly
        embedded_code = (
            "        // Load data (embedded)\n"
            "        allData = " + json.dumps(cleaned_data, ensure_ascii=False) + ";\n"
            "                initializeDashboard();"
        )

        new_html, count = loader_pattern.subn(embedded_code, template_content)
        if count != 1:
            raise ValueError(
                f"Expected exactly 1 fetch loader in dashboard.html, found {count}."
            )

        logger.info(f"Successfully built standalone dashboard with {len(cleaned_data)} records")
        return new_html
    except Exception as e:
        logger.error(f"Error in build_standalone_memory: {str(e)}")
        raise


# ---------------------------------------------------------------------------
# Session-like state stored in app context (fine for single-user local tool)
# For serverless deployment, we store everything in memory
# ---------------------------------------------------------------------------
_state: dict = {
    "step": 1, 
    "record_count": 0, 
    "messages": [], 
    "json_data": None, 
    "standalone_html": None
}


def _render(step=None, messages=None, **kwargs):
    try:
        if step is not None:
            _state["step"] = step
        if messages is not None:
            _state["messages"] = messages
        
        context = {
            "step": _state.get("step", 1),
            "record_count": _state.get("record_count", 0),
            "messages": _state.get("messages", []),
            **kwargs
        }
        
        return render_template_string(PAGE_TEMPLATE, **context)
    except Exception as e:
        logger.error(f"Error in _render: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return f"Template rendering error: {str(e)}", 500


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/", methods=["GET"])
def index():
    try:
        # Reset state for new session
        _state["step"] = 1
        _state["messages"] = []
        _state["json_data"] = None
        _state["standalone_html"] = None
        _state["record_count"] = 0
        
        return _render(step=1, messages=[])
    except Exception as e:
        logger.error(f"Error in index route: {str(e)}")
        return f"Application error: {str(e)}", 500


@app.route("/health")
def health_check():
    """Simple health check endpoint"""
    try:
        import pandas as pd
        import openpyxl
        
        health_info = {
            "status": "healthy",
            "mode": "serverless_memory",
            "dashboard_template_exists": os.path.exists(os.path.join(BASE_DIR, "dashboard.html")),
            "base_dir": BASE_DIR,
            "pandas_version": pd.__version__,
            "openpyxl_available": True,
            "create_template_available": build_template is not None,
            "current_state": {
                "step": _state.get("step", 1),
                "has_data": _state.get("json_data") is not None,
                "has_html": _state.get("standalone_html") is not None,
                "record_count": _state.get("record_count", 0)
            }
        }
        
        # Test if we can create a simple DataFrame
        test_df = pd.DataFrame({"test": [1, 2, 3]})
        health_info["pandas_working"] = len(test_df) == 3
        
        return health_info
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        return {"status": "unhealthy", "error": str(e)}, 500


@app.route("/upload", methods=["POST"])
def upload():
    try:
        logger.debug("Upload route called")
        
        file = request.files.get("excel_file")
        if not file or file.filename == "":
            logger.warning("No file selected")
            return _render(step=1, messages=[("danger", "No file selected.")])

        logger.debug(f"File uploaded: {file.filename}")
        
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if ext not in ALLOWED_EXTENSIONS:
            logger.warning(f"Invalid file extension: {ext}")
            return _render(step=1, messages=[("danger", "Only .xlsx and .xls files are supported.")])

        logger.debug("Processing file in memory (serverless environment)")
        
        try:
            # Process Excel file in memory
            data, count = excel_to_json_memory(file)
            
            # Store data in memory and also save to file for persistence
            _state["json_data"] = data
            _state["record_count"] = count
            
            # Save to jira_data.json file for dashboard fetch requests
            try:
                json_file_path = os.path.join(BASE_DIR, "jira_data.json")
                with open(json_file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                logger.info(f"Successfully saved {count} records to jira_data.json")
            except Exception as json_error:
                logger.warning(f"Failed to save jira_data.json: {json_error}")
                # Continue without failing since data is in memory
            
            logger.info(f"Successfully processed {count} records in memory")
            
            # Show success message with any warnings
            messages = [("success", f"Successfully uploaded and processed {count} records.")]
            
        except ValueError as ve:
            # This is a validation error - show detailed message to user
            error_msg = str(ve)
            if "Excel file validation failed:" in error_msg:
                # Format validation errors nicely for display
                validation_errors = error_msg.replace("Excel file validation failed:\n", "").split("\n")
                formatted_errors = []
                for i, error in enumerate(validation_errors, 1):
                    if error.strip():
                        formatted_errors.append(f"{i}. {error.strip()}")
                
                error_display = "The Excel file has the following issues:\n" + "\n".join(formatted_errors)
                error_display += "\n\nPlease fix these issues and try uploading again. You can download a sample template to see the expected format."
                
                return _render(step=1, messages=[("danger", error_display)])
            else:
                return _render(step=1, messages=[("danger", f"File validation failed: {error_msg}")])
                
        except Exception as e:
            logger.error(f"Conversion failed: {str(e)}")
            return _render(step=1, messages=[("danger", f"File processing failed: {str(e)}. Please ensure your Excel file is not corrupted and follows the expected format.")])

        return _render(step=2, messages=messages)
        
    except Exception as e:
        logger.error(f"Unexpected error in upload route: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return _render(step=1, messages=[("danger", f"Upload failed: {str(e)}")])


@app.route("/generate", methods=["POST"])
def generate():
    try:
        if not _state.get("json_data"):
            return _render(step=1, messages=[("danger", "No data found. Please upload an Excel file first.")])
        
        # Read dashboard template
        template_path = os.path.join(BASE_DIR, "dashboard.html")
        if not os.path.exists(template_path):
            return _render(step=2, messages=[("danger", "dashboard.html template not found in project directory.")])

        with open(template_path, "r", encoding="utf-8") as f:
            template_content = f.read()

        try:
            standalone_html = build_standalone_memory(_state["json_data"], template_content)
            _state["standalone_html"] = standalone_html
            logger.info("Successfully generated standalone dashboard in memory")
        except Exception as e:
            logger.error(f"Dashboard generation failed: {str(e)}")
            return _render(step=2, messages=[("danger", f"Dashboard generation failed: {str(e)}")])

        return _render(step=3, messages=[])
    except Exception as e:
        logger.error(f"Error in generate route: {str(e)}")
        return _render(step=2, messages=[("danger", f"Generation failed: {str(e)}")])


@app.route("/download")
def download():
    try:
        if not _state.get("standalone_html"):
            return redirect(url_for("index"))
        
        # Return the HTML content as a downloadable file
        return Response(
            _state["standalone_html"],
            mimetype="text/html",
            headers={"Content-Disposition": "attachment; filename=dashboard_standalone.html"}
        )
    except Exception as e:
        logger.error(f"Error in download route: {str(e)}")
        return Response(f"Download failed: {str(e)}", status=500)


@app.route("/test-data")
def test_data():
    """Test endpoint to verify data loading"""
    try:
        # Check memory state
        memory_data = _state.get("json_data")
        memory_count = len(memory_data) if memory_data else 0
        
        # Check file
        json_file_path = os.path.join(BASE_DIR, "jira_data.json")
        file_exists = os.path.exists(json_file_path)
        file_count = 0
        file_error = None
        
        if file_exists:
            try:
                with open(json_file_path, 'r', encoding='utf-8') as f:
                    file_data = json.load(f)
                file_count = len(file_data)
            except Exception as e:
                file_error = str(e)
        
        # Create diagnostic HTML
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Data Loading Diagnostic</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
                .container {{ background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .status {{ padding: 15px; margin: 10px 0; border-radius: 5px; }}
                .success {{ background: #d4edda; border-left: 4px solid #28a745; }}
                .error {{ background: #f8d7da; border-left: 4px solid #dc3545; }}
                .info {{ background: #d1ecf1; border-left: 4px solid #17a2b8; }}
                .test-button {{ background: #007bff; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; margin: 10px 5px; }}
                #result {{ margin-top: 20px; padding: 15px; background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>JIRA Dashboard Data Diagnostic</h1>
                
                <h2>Current Status</h2>
                <div class="status {'success' if memory_count > 0 else 'error'}">
                    <strong>Memory State:</strong> {memory_count} records loaded
                </div>
                
                <div class="status {'success' if file_exists else 'error'}">
                    <strong>jira_data.json File:</strong> {'Exists' if file_exists else 'Missing'}
                    {f'({file_count} records)' if file_exists and not file_error else ''}
                    {f'<br>Error: {file_error}' if file_error else ''}
                </div>
                
                <h2>Test Data Loading</h2>
                <button class="test-button" onclick="testJsonEndpoint()">Test /jira_data.json</button>
                <button class="test-button" onclick="testDashboardFetch()">Test Dashboard Fetch</button>
                
                <div id="result"></div>
                
                <h2>Actions</h2>
                <p><a href="/">Upload New Data</a> | <a href="/preview">View Dashboard</a> | <a href="/dashboard">View Dashboard Template</a></p>
                
                <script>
                    async function testJsonEndpoint() {{
                        const result = document.getElementById('result');
                        try {{
                            const response = await fetch('/jira_data.json');
                            const data = await response.json();
                            if (response.ok) {{
                                result.innerHTML = '<div class="status success"><strong>✓ Success:</strong> Loaded ' + (Array.isArray(data) ? data.length : Object.keys(data).length) + ' records from /jira_data.json</div>';
                            }} else {{
                                result.innerHTML = '<div class="status error"><strong>✗ Error:</strong> ' + (data.error || 'Unknown error') + '</div>';
                            }}
                        }} catch (error) {{
                            result.innerHTML = '<div class="status error"><strong>✗ Fetch Error:</strong> ' + error.message + '</div>';
                        }}
                    }}
                    
                    async function testDashboardFetch() {{
                        const result = document.getElementById('result');
                        try {{
                            // Simulate the exact same fetch call as dashboard.html
                            const response = await fetch('jira_data.json');
                            const text = await response.text();
                            const data = JSON.parse(text.replace(/NaN/g, 'null'));
                            result.innerHTML = '<div class="status success"><strong>✓ Dashboard Fetch Success:</strong> Loaded ' + data.length + ' records using dashboard method</div>';
                        }} catch (error) {{
                            result.innerHTML = '<div class="status error"><strong>✗ Dashboard Fetch Error:</strong> ' + error.message + '</div>';
                        }}
                    }}
                </script>
            </div>
        </body>
        </html>
        """
        return html
        
    except Exception as e:
        return f"Diagnostic failed: {str(e)}", 500


@app.route("/jira_data.json")
def serve_jira_data():
    """Serve the JSON data for dashboard consumption"""
    try:
        # First try to get data from memory state
        json_data = _state.get("json_data")
        
        # If not in memory, try to load from file
        if not json_data:
            json_file_path = os.path.join(BASE_DIR, "jira_data.json")
            if os.path.exists(json_file_path):
                try:
                    with open(json_file_path, 'r', encoding='utf-8') as f:
                        json_data = json.load(f)
                    logger.info("Loaded data from jira_data.json file")
                except Exception as file_error:
                    logger.error(f"Failed to load jira_data.json: {file_error}")
        
        if not json_data:
            return {"error": "No data available. Please upload and process an Excel file first."}, 400
        
        # Return the JSON data with proper content type
        return Response(
            json.dumps(json_data, ensure_ascii=False, indent=2),
            mimetype='application/json',
            headers={'Content-Type': 'application/json; charset=utf-8'}
        )
    except Exception as e:
        logger.error(f"Error serving jira_data.json: {str(e)}")
        return {"error": f"Failed to serve data: {str(e)}"}, 500


@app.route("/dashboard")
def serve_dashboard():
    """Serve the dashboard.html template (requires separate jira_data.json fetch)"""
    try:
        # Ensure jira_data.json exists and has data
        json_file_path = os.path.join(BASE_DIR, "jira_data.json")
        if not os.path.exists(json_file_path):
            # Try to create it from memory state if available
            json_data = _state.get("json_data")
            if json_data:
                try:
                    with open(json_file_path, 'w', encoding='utf-8') as f:
                        json.dump(json_data, f, ensure_ascii=False, indent=2)
                    logger.info("Created jira_data.json from memory state")
                except Exception as e:
                    logger.error(f"Failed to create jira_data.json: {e}")
                    return Response("No data file available and failed to create one. Please upload an Excel file first.", status=400)
            else:
                return Response("No data available. Please upload an Excel file first.", status=400)
        
        template_path = os.path.join(BASE_DIR, "dashboard.html")
        if not os.path.exists(template_path):
            return Response("Dashboard template not found", status=404)
        
        with open(template_path, "r", encoding="utf-8") as f:
            dashboard_content = f.read()
        
        return Response(dashboard_content, mimetype="text/html")
    except Exception as e:
        logger.error(f"Error serving dashboard: {str(e)}")
        return Response(f"Dashboard failed to load: {str(e)}", status=500)


@app.route("/preview")
def preview():
    try:
        # DEBUG: Add debug info at the top
        debug_info = []
        
        # First try to get the generated standalone HTML from memory
        standalone_html = _state.get("standalone_html")
        debug_info.append(f"Memory HTML exists: {standalone_html is not None}")
        
        # If not in memory, try to generate it from existing data
        if not standalone_html:
            logger.info("No standalone HTML in memory, attempting to generate from available data")
            debug_info.append("Generating HTML from data...")
            
            # Try to get data from memory first, then from file
            json_data = _state.get("json_data")
            debug_info.append(f"Memory data exists: {json_data is not None}")
            debug_info.append(f"Memory data count: {len(json_data) if json_data else 0}")
            
            if not json_data:
                json_file_path = os.path.join(BASE_DIR, "jira_data.json")
                file_exists = os.path.exists(json_file_path)
                debug_info.append(f"File exists: {file_exists}")
                
                if file_exists:
                    try:
                        with open(json_file_path, 'r', encoding='utf-8') as f:
                            json_data = json.load(f)
                        debug_info.append(f"Loaded {len(json_data)} records from file")
                        logger.info("Loaded data from jira_data.json for preview")
                    except Exception as e:
                        debug_info.append(f"File load error: {str(e)}")
                        logger.error(f"Failed to load jira_data.json: {e}")
                        return Response(f"Debug info: {'; '.join(debug_info)}<br>Failed to load jira_data.json. Error: {str(e)}", status=400, mimetype="text/html")
            
            if json_data:
                # Read dashboard template and generate standalone version
                template_path = os.path.join(BASE_DIR, "dashboard.html")
                template_exists = os.path.exists(template_path)
                debug_info.append(f"Template exists: {template_exists}")
                
                if not template_exists:
                    return Response(f"Debug info: {'; '.join(debug_info)}<br>Dashboard template not found", status=500, mimetype="text/html")
                
                with open(template_path, "r", encoding="utf-8") as f:
                    template_content = f.read()
                
                debug_info.append(f"Template length: {len(template_content)}")
                
                try:
                    standalone_html = build_standalone_memory(json_data, template_content)
                    # Cache it in memory for future requests in this session
                    _state["standalone_html"] = standalone_html
                    debug_info.append("Generated standalone HTML successfully")
                    logger.info("Successfully generated standalone dashboard for preview")
                except Exception as e:
                    debug_info.append(f"Generation error: {str(e)}")
                    logger.error(f"Failed to generate standalone dashboard: {e}")
                    return Response(f"Debug info: {'; '.join(debug_info)}<br>Failed to generate dashboard: {str(e)}", status=500, mimetype="text/html")
            else:
                return Response(f"Debug info: {'; '.join(debug_info)}<br>No data available. Please upload an Excel file first.", status=400, mimetype="text/html")
        
        # Add debug info as HTML comment at the top
        debug_comment = f"<!-- DEBUG: {'; '.join(debug_info)} -->\n"
        standalone_html = debug_comment + standalone_html
        
        return Response(standalone_html, mimetype="text/html")
    except Exception as e:
        logger.error(f"Error in preview route: {str(e)}")
        return Response(f"Preview failed: {str(e)}<br>Debug: Check logs for details", status=500, mimetype="text/html")


@app.route("/template")
def download_template():
    try:
        # Generate template in memory
        if build_template:
            try:
                # Create a temporary in-memory Excel file using the comprehensive template
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import (
                    Alignment,
                    Border,
                    Font,
                    PatternFill,
                    Side,
                )
                from openpyxl.utils import get_column_letter
                from openpyxl.worksheet.datavalidation import DataValidation
                from datetime import date
                
                # Create workbook in memory
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "JIRA Data"
                
                # Define comprehensive headers matching the expected format
                headers = [
                    "Client", "Type", "Project", "Sprint", "EPIC", "Story", "Task",
                    "Quality", "Budget Planned", "Budget Consumed", "Budget Remaining",
                    "Saving Planned", "Saving Achived", "Saving Pending",
                    "Start Date", "Date", "Actual Completion Date", "Story Points", "Priority", "Assignee", "Person", "Status", "Dependencies"
                ]
                
                # Sample realistic data
                sample_rows = [
                    ("Acme Corp", "Story", "Alpha", "Sprint 1", "User Auth", "Login with SSO", 
                     "Frontend implementation", 0, 3000, 1800, 1200, 400, 200, 200, 
                     date(2026, 1, 5), date(2026, 1, 10), date(2026, 1, 8), 3, "High", "Alice", "Alice", "Done", ""),
                    ("Acme Corp", "Task", "Alpha", "Sprint 1", "User Auth", "Login with SSO",
                     "Unit tests", 1, 1200, 900, 300, 150, 90, 60,
                     date(2026, 1, 8), date(2026, 1, 14), date(2026, 1, 16), 2, "High", "Bob", "Bob", "Done", ""),
                    ("Beta Ltd", "Story", "Beta", "Sprint 2", "Reporting", "Management dashboard",
                     "Data model", 0, 5000, 3000, 2000, 600, 300, 300,
                     date(2026, 2, 5), date(2026, 2, 11), None, 8, "Medium", "Eve", "Eve", "In Progress", ""),
                ]
                
                # Styles
                header_fill = PatternFill("solid", fgColor="1F6FEB")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                thin = Side(style="thin", color="D0D7DE")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                
                # Add headers
                ws.row_dimensions[1].height = 32
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = border
                
                # Add sample data
                for row_idx, row_data in enumerate(sample_rows, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(vertical="center")
                        
                        # Format dates
                        if isinstance(value, date):
                            cell.number_format = "YYYY-MM-DD"
                        elif col_idx in [8, 9, 10, 11, 12, 13, 14, 18]:  # Numeric columns (updated indices for new column positions)
                            cell.number_format = "#,##0"
                
                # Set column widths
                col_widths = {
                    "Client": 15, "Type": 10, "Project": 10, "Sprint": 12, "EPIC": 20,
                    "Story": 36, "Task": 30, "Quality": 9,
                    "Budget Planned": 15, "Budget Consumed": 16, "Budget Remaining": 17,
                    "Saving Planned": 15, "Saving Achived": 15, "Saving Pending": 14,
                    "Start Date": 13, "Date": 13, "Actual Completion Date": 18, "Story Points": 13,
                    "Priority": 10, "Assignee": 12, "Person": 12, "Status": 13, "Dependencies": 16,
                }
                
                for col_idx, header in enumerate(headers, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 14)
                
                # Add data validation dropdowns
                last_row = 200  # Allow for many future rows
                
                # Type validation (now column B, was column A)
                type_validation = DataValidation(
                    type="list",
                    formula1='"Story,Task,Bug,Epic,Subtask"',
                    allow_blank=True,
                    showDropDown=True
                )
                type_validation.sqref = f"B2:B{last_row}"
                ws.add_data_validation(type_validation)
                
                # Priority validation (now column S, was column Q)
                priority_validation = DataValidation(
                    type="list",
                    formula1='"Low,Medium,High,Critical,Normal"',
                    allow_blank=True,
                    showDropDown=True
                )
                priority_validation.sqref = f"S2:S{last_row}"  # Updated column position for Priority
                ws.add_data_validation(priority_validation)
                
                # Status validation
                status_validation = DataValidation(
                    type="list", 
                    formula1='"To Do,In Progress,Done,Blocked,Review"',
                    allow_blank=True,
                    showDropDown=True
                )
                status_validation.sqref = f"V2:V{last_row}"  # Updated column position for Status
                ws.add_data_validation(status_validation)
                
                # Freeze header row and add auto-filter
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
                
                # Add Instructions sheet
                ws2 = wb.create_sheet("Instructions")
                ws2.column_dimensions["A"].width = 90
                
                instructions = [
                    ("JIRA Dashboard Template — Instructions", True),
                    ("", False),
                    ("Required columns (do NOT rename or remove these):", True),
                    ("  Client          — Client or customer name (e.g. Acme Corp, Beta Ltd)", False),
                    ("  Type            — Ticket type: Story, Task, Bug, Epic, Subtask", False),
                    ("  Project         — Project key or name (e.g. Alpha, Beta)", False),
                    ("  Sprint          — Sprint name (e.g. Sprint 1)", False),
                    ("  EPIC            — Epic name the ticket belongs to", False),
                    ("  Story           — Parent story description", False),
                    ("  Task            — Task description", False),
                    ("  Quality         — Integer bug count (0 = no bugs)", False),
                    ("  Budget Planned  — Planned effort in story points", False),
                    ("  Budget Consumed — Story points consumed so far", False),
                    ("  Budget Remaining— Remaining story points", False),
                    ("  Saving Planned  — Planned monetary savings", False),
                    ("  Saving Achived  — Achieved monetary savings (note: original spelling)", False),
                    ("  Saving Pending  — Pending monetary savings", False),
                    ("  Start Date      — Task start date in YYYY-MM-DD format", False),
                    ("  Date            — Planned completion date (YYYY-MM-DD format)", False),
                    ("  Actual Completion Date — Actual completion date for Done tasks (YYYY-MM-DD)", False),
                    ("  Story Points    — Effort estimate (integer)", False),
                    ("  Priority        — High / Medium / Low / Critical", False),
                    ("  Assignee        — Person responsible", False),
                    ("  Person          — Person name (can be same as Assignee)", False),
                    ("  Status          — To Do / In Progress / Done / Blocked / Review", False),
                    ("  Dependencies    — Optional free-text dependency notes", False),
                    ("", False),
                    ("Validation Rules:", True),
                    ("  • All numeric columns must contain valid numbers", False),
                    ("  • Dates must be in a standard format (YYYY-MM-DD recommended)", False),
                    ("  • Actual Completion Date: Only fill for tasks with Status = 'Done'", False),
                    ("  • Priority must be: Low, Medium, High, or Critical", False),
                    ("  • Status must be: To Do, In Progress, Done, Blocked, or Review", False),
                    ("  • Type must be: Story, Task, Bug, Epic, or Subtask", False),
                    ("  • Story Points must be positive integers", False),
                    ("  • Budget columns represent story points (effort), not money", False),
                    ("  • Saving columns represent monetary values", False),
                    ("  • Quality = bug count (0 = no bugs, higher = more bugs)", False),
                    ("", False),
                    ("Key Features in Dashboard:", True),
                    ("  • Budget fields displayed as story points (e.g., '150 pts')", False),
                    ("  • Completed tasks show both planned and actual completion dates", False),
                    ("  • Late deliveries highlighted in red, on-time in green", False),
                    ("  • User performance metrics: efficiency, on-time %, bug rate", False),
                    ("  • Three-level health indicators: Efforts, Timeline, Quality", False),
                ]
                
                title_font = Font(bold=True, size=12, color="1F6FEB")
                bold_font = Font(bold=True, size=10)
                norm_font = Font(size=10)
                
                for r_idx, (text, is_header) in enumerate(instructions, 1):
                    cell = ws2.cell(row=r_idx, column=1, value=text)
                    if r_idx == 1:
                        cell.font = title_font
                    elif is_header:
                        cell.font = bold_font
                    else:
                        cell.font = norm_font
                
                # Save to BytesIO
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                return Response(
                    excel_buffer.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=JIRA_template.xlsx"}
                )
                
            except Exception as e:
                logger.error(f"Failed to generate template in memory: {e}")
                return Response(f"Could not generate template: {str(e)}", status=500, mimetype="text/plain")
        else:
            return Response("Template generation not available", status=500, mimetype="text/plain")
            
    except Exception as e:
        logger.error(f"Error in template download: {str(e)}")
        return Response(f"Template download failed: {str(e)}", status=500, mimetype="text/plain")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {error}")
    return "Internal server error occurred. Please check the logs for details.", 500

@app.errorhandler(Exception)
def handle_exception(e):
    logger.error(f"Unhandled exception: {str(e)}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    return f"An error occurred: {str(e)}", 500

if __name__ == "__main__":
    print("Starting JIRA Dashboard Generator at http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
