"""
Generate JIRA_template.xlsx — a sample Excel template with realistic data,
column headers, data-validation dropdowns, and formatting hints.

Run once:  python create_template.py
The Flask app calls this automatically on first startup if the file is missing.
"""

import os
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Sample rows (10 realistic tickets)
# ---------------------------------------------------------------------------
SAMPLE_ROWS = [
    # Client, Type,    Project,  Sprint,    EPIC,            Story,                          Task,                    Quality, BudgetPlanned, BudgetConsumed, BudgetRemaining, SavingPlanned, SavingAchieved, SavingPending, StartDate,       Date,       StoryPoints, Priority, Assignee, Status,      Dependencies
    ("Acme Corp", "Story",  "Alpha",  "Sprint 1","User Auth",     "Login with SSO",               "Frontend implementation",  0,  3000, 1800, 1200,  400, 200, 200, date(2026, 1, 5), date(2026, 1, 10), 3, "High",   "Alice",   "Done",        ""),
    ("Acme Corp", "Task",   "Alpha",  "Sprint 1","User Auth",     "Login with SSO",               "Unit tests",               1,  1200,  900,  300,  150,  90,  60, date(2026, 1, 8), date(2026, 1, 14), 2, "High",   "Bob",     "Done",        ""),
    ("Acme Corp", "Story",  "Alpha",  "Sprint 2","User Auth",     "Password reset flow",          "API endpoint",             0,  2500, 2500,    0,  300, 300,   0, date(2026, 1, 20), date(2026, 1, 28), 5, "Medium", "Carol",   "Done",        ""),
    ("Beta Ltd",  "Task",   "Alpha",  "Sprint 2","User Auth",     "Password reset flow",          "Email template design",    0,   800,  400,  400,  100,  50,  50, date(2026, 1, 25), date(2026, 2,  4), 1, "Low",    "Dave",    "In Progress", ""),
    ("Beta Ltd",  "Story",  "Beta",   "Sprint 3","Reporting",     "Management dashboard",         "Data model",               0,  5000, 3000, 2000,  600, 300, 300, date(2026, 2, 5), date(2026, 2, 11), 8, "High",   "Eve",     "In Progress", ""),
    ("Gamma Inc", "Task",   "Beta",   "Sprint 3","Reporting",     "Management dashboard",         "Chart components",         2,  2200, 1100, 1100,  250, 100, 150, date(2026, 2, 10), date(2026, 2, 18), 3, "Medium", "Frank",   "In Progress", ""),
    ("Gamma Inc", "Bug",    "Beta",   "Sprint 3","Reporting",     "Management dashboard",         "Fix date filter",          3,   600,    0,  600,    0,   0,   0, date(2026, 2, 15), date(2026, 2, 20), 1, "High",   "Grace",   "To Do",       ""),
    ("Delta Co",  "Story",  "Beta",   "Sprint 4","Integrations",  "Connect to third-party API",   "OAuth handshake",          0,  4000, 4000,    0,  500, 500,   0, date(2026, 2, 28), date(2026, 3,  5), 5, "High",   "Henry",   "Done",        ""),
    ("Delta Co",  "Task",   "Gamma",  "Sprint 4","Integrations",  "Connect to third-party API",   "Rate-limit handling",      1,  1500,  750,  750,  200,  80, 120, date(2026, 3, 5), date(2026, 3, 12), 2, "Medium", "Iris",    "In Progress", ""),
    ("Epsilon LLC", "Story",  "Gamma",  "Sprint 5","Performance",   "Reduce API response time",     "Query optimisation",       0,  3500,  500, 3000,  450,  50, 400, date(2026, 3, 10), date(2026, 3, 19), 5, "Medium", "Jack",    "To Do",       ""),
]

COLUMNS = [
    "Client", "Type", "Project", "Sprint", "EPIC", "Story", "Task",
    "Quality", "Budget Planned", "Budget Consumed", "Budget Remaining",
    "Saving Planned", "Saving Achived", "Saving Pending",
    "Start Date", "Date", "Story Points", "Priority", "Assignee", "Status", "Dependencies",
]

# ---------------------------------------------------------------------------
# Validation lists (must match values used in dashboard.html filters)
# ---------------------------------------------------------------------------
TYPES      = ["Story", "Task", "Bug", "Sub-task", "Epic"]
PRIORITIES = ["High", "Medium", "Low", "Critical"]
STATUSES   = ["To Do", "In Progress", "Done", "Blocked", "In Review"]


def build_template(output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JIRA Data"

    # ── Styles ──────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    even_fill = PatternFill("solid", fgColor="F6F8FA")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    num_fmt   = "#,##0"
    date_fmt  = "YYYY-MM-DD"

    # ── Header row ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border

    # ── Data rows ───────────────────────────────────────────────────────────
    for row_idx, row in enumerate(SAMPLE_ROWS, start=2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            col_name = COLUMNS[col_idx - 1]
            if col_name in (
                "Budget Planned", "Budget Consumed", "Budget Remaining",
                "Saving Planned", "Saving Achived", "Saving Pending",
                "Quality", "Story Points",
            ):
                cell.number_format = num_fmt
            elif col_name in ("Date", "Start Date"):
                cell.number_format = date_fmt

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = {
        "Client": 15, "Type": 10, "Project": 10, "Sprint": 12, "EPIC": 20,
        "Story": 36, "Task": 30, "Quality": 9,
        "Budget Planned": 15, "Budget Consumed": 16, "Budget Remaining": 17,
        "Saving Planned": 15, "Saving Achived": 15, "Saving Pending": 14,
        "Start Date": 13, "Date": 13, "Story Points": 13,
        "Priority": 10, "Assignee": 12, "Status": 13, "Dependencies": 16,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    # ── Data validations ─────────────────────────────────────────────────────
    last_row = len(SAMPLE_ROWS) + 200   # cover plenty of future rows

    def _col_letter(name):
        return get_column_letter(COLUMNS.index(name) + 1)

    def _dv(formula, col_name):
        col = _col_letter(col_name)
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            sqref=f"{col}2:{col}{last_row}",
        )
        ws.add_data_validation(dv)

    _dv('"' + ",".join(TYPES)      + '"', "Type")
    _dv('"' + ",".join(PRIORITIES) + '"', "Priority")
    _dv('"' + ",".join(STATUSES)   + '"', "Status")

    # ── Freeze header & add auto-filter ──────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Instructions sheet ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 90

    instructions = [
        ("JIRA Dashboard Template — Instructions", True),
        ("", False),
        ("Required columns (do NOT rename or remove these):", True),
        ("  Client          — Client or customer name (e.g. Acme Corp, Beta Ltd)", False),
        ("  Type            — Ticket type: Story, Task, Bug, Sub-task, Epic", False),
        ("  Project         — Project key or name (e.g. Alpha, Beta)", False),
        ("  Sprint          — Sprint name (e.g. Sprint 1)", False),
        ("  EPIC            — Epic name the ticket belongs to", False),
        ("  Story           — Parent story description", False),
        ("  Task            — Task description", False),
        ("  Quality         — Integer quality score (0 = no issues)", False),
        ("  Budget Planned  — Planned budget in your currency", False),
        ("  Budget Consumed — Budget spent so far", False),
        ("  Budget Remaining— Remaining budget (can be formula = Planned - Consumed)", False),
        ("  Saving Planned  — Planned savings", False),
        ("  Saving Achived  — Achieved savings (note: keep original spelling)", False),
        ("  Saving Pending  — Pending savings", False),
        ("  Start Date      — Task start date in YYYY-MM-DD format", False),
        ("  Date            — Task target/due date in YYYY-MM-DD format", False),
        ("  Story Points    — Effort estimate (integer)", False),
        ("  Priority        — High / Medium / Low / Critical", False),
        ("  Assignee        — Person responsible", False),
        ("  Status          — To Do / In Progress / Done / Blocked / In Review", False),
        ("  Dependencies    — Optional free-text dependency notes", False),
        ("", False),
        ("Tips:", True),
        ("  • Keep column order exactly as shown in the JIRA Data sheet.", False),
        ("  • Dates must be parseable by pandas (YYYY-MM-DD recommended).", False),
        ("  • The 'Cost' column is ignored if present — safe to include or omit.", False),
        ("  • Remove these instruction rows before uploading.", False),
    ]

    title_font = Font(bold=True, size=12, color="1F6FEB")
    bold_font  = Font(bold=True, size=10)
    norm_font  = Font(size=10)

    for r_idx, (text, is_header) in enumerate(instructions, start=1):
        cell = ws2.cell(row=r_idx, column=1, value=text)
        if r_idx == 1:
            cell.font = title_font
        elif is_header:
            cell.font = bold_font
        else:
            cell.font = norm_font

    wb.save(output_path)
    print(f"Template saved → {output_path}")


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "JIRA_template.xlsx")
    build_template(out)
